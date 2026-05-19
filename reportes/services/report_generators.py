from abc import ABC, abstractmethod
from io import BytesIO

from django.utils.translation import gettext as _

from facturas.models import Factura
from proveedores.models import Proveedor
from conciliacion.models import ProcesoReconciliacion
from core.services.pdf_generator import (
    generate_facturas_report,
    generate_conciliacion_report,
    generate_proveedores_report,
)


class ReportGenerator(ABC):
    @abstractmethod
    def generate(self, data):
        ...


def _facturas_queryset(fecha_desde=None, fecha_hasta=None):
    qs = Factura.objects.select_related('proveedor').all()
    if fecha_desde:
        qs = qs.filter(fecha_emision__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha_emision__lte=fecha_hasta)
    return qs.order_by('-fecha_emision')


class PDFReportGenerator(ReportGenerator):
    mime_type = 'application/pdf'

    def generate(self, data):
        data = data or {}
        tipo = data.get('tipo', 'facturas')
        fecha_desde = data.get('fecha_desde')
        fecha_hasta = data.get('fecha_hasta')

        if tipo == 'facturas':
            buffer = generate_facturas_report(_facturas_queryset(fecha_desde, fecha_hasta))
        elif tipo == 'conciliacion':
            proceso = ProcesoReconciliacion.objects.order_by('-created_at').first()
            buffer = generate_conciliacion_report(proceso)
        else:
            buffer = generate_proveedores_report(Proveedor.objects.all())

        filename = _("reporte_%(tipo)s.pdf") % {"tipo": tipo}
        return buffer.getvalue(), filename, self.mime_type


class ExcelReportGenerator(ReportGenerator):
    mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def generate(self, data):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        data = data or {}
        tipo = data.get('tipo', 'facturas')
        fecha_desde = data.get('fecha_desde')
        fecha_hasta = data.get('fecha_hasta')

        wb = Workbook()
        ws = wb.active

        if tipo == 'facturas':
            ws.title = 'Facturas'
            headers = [_('Número'), _('Proveedor'), _('Monto Total'), _('Monto Pagado'), _('Estado'), _('Fecha emisión')]
            rows = [
                [f.numero_factura, f.proveedor.nombre,
                 float(f.monto_total), float(f.monto_pagado),
                 f.get_estado_display(),
                 f.fecha_emision.isoformat() if f.fecha_emision else '']
                for f in _facturas_queryset(fecha_desde, fecha_hasta)
            ]
        elif tipo == 'conciliacion':
            ws.title = 'Conciliacion'
            proceso = ProcesoReconciliacion.objects.order_by('-created_at').first()
            headers = [_('Factura'), _('Proveedor'), _('Monto'), _('Tipo'), _('Confianza %')]
            rows = []
            if proceso:
                for c in proceso.coincidencia_set.select_related('factura', 'factura__proveedor'):
                    rows.append([
                        c.factura.numero_factura,
                        c.factura.proveedor.nombre,
                        float(c.factura.monto_total),
                        c.get_tipo_display(),
                        float(c.porcentaje_confianza),
                    ])
        else:
            ws.title = 'Proveedores'
            headers = [_('Nombre'), _('NIT'), _('Email'), _('Teléfono'), _('Facturas')]
            rows = [
                [p.nombre, p.nit, p.email or '', p.telefono or '', p.facturas.count()]
                for p in Proveedor.objects.all()
            ]

        ws.append(headers)
        header_fill = PatternFill(start_color='FF1E293B', end_color='FF1E293B', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFFFF')
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='left')

        for row in rows:
            ws.append(row)

        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(15, len(str(header)) + 4)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = _("reporte_%(tipo)s.xlsx") % {"tipo": tipo}
        return buffer.getvalue(), filename, self.mime_type
